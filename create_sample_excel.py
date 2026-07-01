"""Create a sample emails.xlsx file to show the expected structure."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Applications"

# headers
headers = ["email", "company", "position", "status", "notes"]
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# sample data
sample_data = [
    ["hr@techcorp.com", "TechCorp", "Software Engineer", "", "Found on LinkedIn"],
    ["jobs@bigagency.com", "Big Agency", "Frontend Developer", "", "Company website"],
    ["recruiter@spammy.com", "Spammy Inc", "Full Stack Dev", "spam", "Marked as spam — do not send"],
    ["talent@startup.io", "Startup.io", "Backend Engineer", "", "Referral from Ahmed"],
    ["blocked@blocked.com", "Blocked Corp", "DevOps Engineer", "blocked", "Bounced last time"],
    ["careers@enterprise.com", "Enterprise Ltd", "Python Developer", "", "Applied via referral"],
    ["already@sent.com", "Sent Inc", "Data Engineer", "sent", "Already sent on 2025-01-15"],
]

# status colour fills
status_fills = {
    "spam": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "blocked": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "sent": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "replied": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
}

for row_idx, row_data in enumerate(sample_data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")
        # colour the row based on status (column 4)
        status = str(row_data[3]).lower() if row_data[3] else ""
        if status in status_fills:
            cell.fill = status_fills[status]

# column widths
ws.column_dimensions["A"].width = 30
ws.column_dimensions["B"].width = 20
ws.column_dimensions["C"].width = 25
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 35

# freeze header row
ws.freeze_panes = "A2"

# add a legend sheet
legend = wb.create_sheet("Legend")
legend_data = [
    ("Status", "Meaning"),
    ("", "New — ready to send"),
    ("spam", "Recipient marked us as spam — skip"),
    ("blocked", "Email bounced or is blocked — skip"),
    ("sent", "Already sent — skip"),
    ("replied", "They replied — skip"),
]
for r, (status, meaning) in enumerate(legend_data, start=1):
    c1 = legend.cell(row=r, column=1, value=status)
    c2 = legend.cell(row=r, column=2, value=meaning)
    if r == 1:
        c1.font = Font(bold=True)
        c2.font = Font(bold=True)
    c1.border = thin_border
    c2.border = thin_border
    if status in status_fills:
        c1.fill = status_fills[status]
        c2.fill = status_fills[status]

legend.column_dimensions["A"].width = 15
legend.column_dimensions["B"].width = 40

wb.save("emails.xlsx")
print("Created emails.xlsx with sample data.")
