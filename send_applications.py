"""
Job Application Email Sender
Reads an Excel sheet of recipient emails and sends personalized job application
emails via Gmail SMTP. Skips any addresses that have been marked as blocked or
detected as spam. Includes rate limiting to avoid triggering spam filters.
"""

import smtplib
import ssl
import time
import random
import configparser
import argparse
import sys
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)


# ── helpers ──────────────────────────────────────────────────────────────────

def load_config(config_path: str) -> dict:
    """Read settings from an INI-style config file."""
    cfg = configparser.ConfigParser()
    cfg.read(config_path)
    if "gmail" not in cfg:
        raise ValueError(f"[gmail] section not found in {config_path}")
    g = cfg["gmail"]
    required = ["sender_email", "app_password"]
    for key in required:
        if key not in g:
            raise ValueError(f"Missing required key '{key}' in [gmail] section")
            
    def get_clean(key, default=""):
        val = g.get(key, default)
        if val is not None:
            val = str(val).strip().strip("'\"").strip()
        return val
        
    return {
        "sender_email": get_clean("sender_email"),
        "app_password": get_clean("app_password"),
        "sender_name": get_clean("sender_name", g["sender_email"]),
        "resume_path": get_clean("resume_path"),
        "cover_letter_path": get_clean("cover_letter_path"),
        "min_delay": float(get_clean("min_delay", "30")),
        "max_delay": float(get_clean("max_delay", "90")),
        "default_position": get_clean("default_position"),
    }


def load_workbook_data(xlsx_path: str) -> list[dict]:
    """Load rows from Excel workbook or CSV file."""
    path = Path(xlsx_path)
    if path.suffix.lower() == ".csv":
        import csv
        try:
            with open(path, mode="r", encoding="utf-8-sig") as f:
                rows = list(csv.reader(f))
        except UnicodeDecodeError:
            with open(path, mode="r", encoding="cp1256", errors="ignore") as f:
                rows = list(csv.reader(f))
    else:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()

    if not rows:
        raise ValueError("File is empty.")

    # Detect if the first row is a header row
    first_row = rows[0]
    has_header = False
    
    # Check if first cell of first row contains '@' (email address)
    first_cell = str(first_row[0]).strip().lower() if first_row[0] is not None else ""
    
    header_keywords = {
        "email", "email address", "e-mail", "mail", "البريد الإلكتروني", "البريد الالكتروني", "الايميل",
        "position", "job title", "job_title", "title", "job", "role", "المسمى الوظيفي", "الوظيفة",
        "company", "company name", "company_name", "organization", "firm", "اسم الشركة", "الشركة"
    }
    
    if first_cell and "@" not in first_cell:
        # If the first cell matches any header keyword, or if it doesn't look like an email
        if first_cell in header_keywords or not any(domain in first_cell for domain in (".com", ".net", ".org", ".edu", ".gov", ".co", ".sa")):
            has_header = True

    header_mapping = {}
    data_start_idx = 1
    
    if has_header:
        raw_headers = [str(h).strip().lower() if h else f"col_{i}" for i, h in enumerate(first_row)]
        for idx, h in enumerate(raw_headers):
            if h in ("email", "email address", "e-mail", "mail", "البريد الإلكتروني", "البريد الالكتروني", "الايميل"):
                header_mapping[idx] = "email"
            elif h in ("position", "job title", "job_title", "title", "job", "role", "المسمى الوظيفي", "الوظيفة"):
                header_mapping[idx] = "position"
            elif h in ("company", "company name", "company_name", "organization", "firm", "اسم الشركة", "الشركة"):
                header_mapping[idx] = "company"
            elif h in ("status", "state", "الحالة"):
                header_mapping[idx] = "status"
            elif h in ("notes", "note", "comment", "ملاحظات"):
                header_mapping[idx] = "notes"
            else:
                header_mapping[idx] = h
    else:
        # No header row. Map by index order: Column 1 is email, Column 2 is position, etc.
        data_start_idx = 0
        header_mapping = {
            0: "email",
            1: "position",
            2: "company",
            3: "status",
            4: "notes"
        }

    records = []
    for row in rows[data_start_idx:]:
        record = {
            "email": "",
            "company": "",
            "position": "",
            "status": "",
            "notes": ""
        }
        for idx, val in enumerate(row):
            if idx in header_mapping:
                key = header_mapping[idx]
                if val is not None:
                    record[key] = str(val).strip()
        
        # Only add records that have at least some data (like an email)
        if record["email"]:
            records.append(record)
            
    return records


def should_skip(record: dict) -> bool:
    """Return True if this recipient should be skipped."""
    status = record.get("status", "").lower()
    skip_values = {"blocked", "spam", "bounced", "skip", "sent", "replied", "ignore"}
    if status in skip_values:
        return True
    email = record.get("email", "")
    if not email or "@" not in email:
        return True
    return False


def build_email(
    sender_email: str,
    sender_name: str,
    recipient_email: str,
    company: str,
    position: str,
    body_template: str,
) -> MIMEMultipart:
    """Compose a MIME email with the given template filled in."""
    msg = MIMEMultipart()
    msg["From"] = f"{sender_name} <{sender_email}>"
    msg["To"] = recipient_email
    
    company = (company or "").strip()
    position = (position or "").strip()
    
    if company:
        msg["Subject"] = f"Application for {position} at {company}"
    else:
        msg["Subject"] = f"Application for {position}"

    # Prepare values for formatting the body
    salutation = f"Dear Hiring Team at {company}" if company else "Dear Hiring Team"
    company_at = f" at {company}" if company else ""

    body = body_template.format(
        salutation=salutation,
        company_at=company_at,
        position=position,
        recipient_email=recipient_email,
    )
    msg.attach(MIMEText(body, "plain"))
    return msg


def attach_file(msg: MIMEMultipart, file_path: str):
    """Attach a file to the MIME message."""
    path = Path(file_path)
    if not path.exists():
        print(f"  WARNING: attachment not found — {file_path}")
        return
    with open(path, "rb") as f:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(f.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename={path.name}")
    msg.attach(part)


def send_email(
    smtp_server: smtplib.SMTP_SSL,
    sender_email: str,
    recipient_email: str,
    msg: MIMEMultipart,
) -> bool:
    """Send a single email. Returns True on success."""
    try:
        smtp_server.sendmail(sender_email, recipient_email, msg.as_string())
        return True
    except smtplib.SMTPRecipientsRefused:
        print(f"  RECIPIENT REFUSED (likely blocked/spam): {recipient_email}")
        return False
    except smtplib.SMTPException as exc:
        print(f"  SMTP error sending to {recipient_email}: {exc}")
        return False


# ── main ─────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Send job application emails from an Excel or CSV list.")
    parser.add_argument("--xlsx", default="emails.xlsx", help="Path to the Excel (.xlsx) or CSV (.csv) file (default: emails.xlsx)")
    parser.add_argument("--config", default="config.ini", help="Path to the config file (default: config.ini)")
    parser.add_argument("--dry-run", action="store_true", help="Preview emails without sending")
    parser.add_argument("--limit", type=int, default=0, help="Max number of emails to send (0 = no limit)")
    parser.add_argument("--position", default="", help="Default job title/position to apply for")
    args = parser.parse_args()

    # load config
    config = load_config(args.config)
    sender_email = config["sender_email"]
    app_password = config["app_password"]
    sender_name = config["sender_name"]
    resume_path = config["resume_path"]
    cover_letter_path = config["cover_letter_path"]
    min_delay = config["min_delay"]
    max_delay = config["max_delay"]
    default_position = config.get("default_position", "")

    # load recipients
    records = load_workbook_data(args.xlsx)
    total = len(records)
    print(f"Loaded {total} rows from {args.xlsx}")

    # email body template — edit this to match your style
    body_template = """{salutation},

I am writing to express my interest in the {position} position{company_at}. I believe my skills and experience make me a strong candidate for this role.

I have attached my resume. I would welcome the opportunity to discuss how I can contribute to your team.

Thank you for your time and consideration.

Best regards,
{sender_name}
""".replace("{sender_name}", sender_name)

    # connect to Gmail SMTP
    if not args.dry_run:
        context = ssl.create_default_context()
        print("Connecting to Gmail SMTP server...")
        try:
            smtp_server = smtplib.SMTP_SSL("smtp.gmail.com", 465, context=context)
            smtp_server.login(sender_email, app_password)
        except smtplib.SMTPAuthenticationError:
            print("Authentication failed. Check your email and App Password in config.ini")
            sys.exit(1)
        except Exception as exc:
            print(f"Could not connect to Gmail: {exc}")
            sys.exit(1)
        print("Connected and authenticated.\n")

    sent = 0
    skipped = 0
    failed = 0

    for i, record in enumerate(records, start=1):
        email = record.get("email", "")
        company = record.get("company", "")
        position = record.get("position", "") or args.position or default_position or "open position"

        comp_display = company if company else "(no company)"
        print(f"[{i}/{total}] {email} — {comp_display} ({position})")

        if should_skip(record):
            reason = record.get("status", "")
            if not reason:
                if not email or "@" not in email:
                    reason = "missing or invalid email"
                else:
                    reason = "skipped"
            print(f"  SKIPPED (status: {reason})")
            skipped += 1
            continue

        # build message
        msg = build_email(sender_email, sender_name, email, company, position, body_template)

        if resume_path:
            attach_file(msg, resume_path)
        if cover_letter_path:
            attach_file(msg, cover_letter_path)

        if args.dry_run:
            if sent == 0:
                print(f"\n--- PREVIEW OF THE FIRST EMAIL ---")
                print(f"Subject: {msg['Subject']}")
                # Extract body
                body_part = ""
                for part in msg.walk():
                    if part.get_content_type() == "text/plain":
                        body_part = part.get_payload()
                        break
                print("Body:")
                print(body_part)
                attachments_list = [Path(p).name for p in [resume_path, cover_letter_path] if p]
                print(f"Attachments: {attachments_list}")
                print(f"----------------------------------\n")
            print(f"  DRY RUN — would send to {email}")
            sent += 1
            continue

        # send
        success = send_email(smtp_server, sender_email, email, msg)
        if success:
            print(f"  SENT successfully")
            sent += 1
        else:
            print(f"  FAILED — will need to retry or mark as blocked")
            failed += 1

        # rate limiting delay (skip after the last email)
        if i < total and (args.limit == 0 or sent < args.limit):
            delay = random.uniform(min_delay, max_delay)
            print(f"  Waiting {delay:.0f}s before next email...")
            time.sleep(delay)

        # honour --limit
        if args.limit and sent >= args.limit:
            print(f"\nReached limit of {args.limit} emails. Stopping.")
            break

    if not args.dry_run:
        smtp_server.quit()

    print(f"\n{'='*50}")
    print(f"Summary: {sent} sent, {skipped} skipped, {failed} failed (out of {total} total)")


if __name__ == "__main__":
    main()
